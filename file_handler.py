from win32com import client
import openpyxl as excel
import os
import win32com


class FileHandler:

    def __init__(self) -> None:
        self.__path = f'C:\\Users\\fabio\\OneDrive\\Documentos\\My projects\\Musical Instruments Repair Shop System\\'
        self.__file = 'work_order.xlsx'

    def read_file(self):

        file = excel.load_workbook(f"{self.__path}files\\{self.__file}")
        sheet = file.active
        for items in sheet.values:
            print(items)

    def add_info_to_excel(self, id, name, phone, email,order_number, services: list):

        file = excel.load_workbook(f"{self.__path}files\\{self.__file}")
        sheet = file.active

        sheet.cell(row=6, column=7).value = f"#00{order_number}"
        sheet.cell(row=7, column=7).value = id
        sheet.cell(row=10, column=2).value = name
        sheet.cell(row=11, column=2).value = email
        sheet.cell(row=12, column=2).value = phone
        

        # row_number = 18 and column_number = 3 are the beginning of the sheet to insert services details

        for i in range(len(services)):
            sheet.cell(row=18+i, column=2).value = 1+i
            sheet.cell(row=18+i, column=3).value = services[i].get_services_details()[0]
            sheet.cell(row=18+i, column=4).value = services[i].get_services_details()[1]
            sheet.cell(row=18+i, column=6).value = services[i].get_services_details()[2]

        self.save_invoice(id, order_number, services)

        file.save(f"{self.__path}files\\{name}_{self.__file}")   

    def convert_to_pdf(self, name):

        app = client.DispatchEx("Excel.Application")
        app.Interactive = False
        app.Visible = False

        wb = app.Workbooks.open(f'{self.__path}files\\{name}_{self.__file}')
        # output = os.path.splitext(file_location)[0]
        output = f'{self.__path}work_orders\\{name}_work_order'      
        wb.ActiveSheet.ExportAsFixedFormat(0, output)
        wb.Close()
        os.remove(f'{self.__path}files\\{name}_{self.__file}')
    
    def save_invoice(self, id, order_number, services):
 
        invoice_string = f"{id};{order_number};"

        for service in services:
            invoice_string += f"{service.get_services_details()};"

        invoice_string = invoice_string[:-1]

        with open('./files/invoices.csv', 'a') as invoices:
            invoices.write(f"{invoice_string}\n")
        
        invoices.close()
        del invoice_string

    def store_client(self, client_details: tuple):

        client_string = ""
        for detail in client_details:
            client_string += f"{detail};"
        client_string = client_string[:-1]

        with open('./files/clients.csv', 'a') as clients:
            clients.write(f"{client_string}\n")
        clients.close()
        del client_string

    def load_clients(self):

        with open('./files/clients.csv') as clients_file:

            list_of_clients = []
            for client in clients_file:
                client = client.strip()
                splitted = client.split(';')
                list_of_clients.append(splitted)

        clients_file.close()

        return list_of_clients
    
    def load_services(self):

        list_of_services = []
        with open('./files/invoices.csv') as invoices_file:
            for services in invoices_file:
                services = services.strip()
                splitted = services.split(';')
                list_of_services.append(splitted)

        invoices_file.close()
        return list_of_services